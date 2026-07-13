"""
Syncs Job data from monthly "Material Purchase Closing Report" Excel files
into the Airtable "Jobs" table.

Usage examples:
    # 1) Set the Airtable token as an environment variable
    export AIRTABLE_TOKEN="patXXXXXXXXXXXXXX"

    # 2) Process every "Material Purchase Closing Report (YYYY-MM)" file
    #    in a folder, newest month -> oldest month (dry-run recommended first)
    python3 sync_jobs.py --folder ./data --dry-run
    python3 sync_jobs.py --folder ./data

    # Or target specific files only
    python3 sync_jobs.py --files "2026-06.xlsx" "2026-05.xlsx"

How it works:
  - Extracts (YYYY, MM) from the filename. Works with any year (2024, 2025,
    2026, ...) since the pattern just looks for a 4-digit "20xx" year followed
    by a 1-2 digit month, e.g. "Material Purchase Closing Report
    (2026-04).xlsx" or the underscore-sanitized "Material_Purchase_
    Closing_Report__2026-04_.xlsx" both match.
  - Within each file, finds the "previous month, confirmed" sheet (e.g. the
    2026-04 file -> sheet "26 3월확정"). The "추정" (estimate) sheet for the
    current month is intentionally excluded.
  - Data starts at row 13. Column A = Business Unit, Column B = Job Code
    (PJT CODE), Column C = Job Name (PROJECT NAME).
  - Stops reading a sheet as soon as it hits a row where both A and B are
    empty (this naturally excludes the "상품매출"/UC00001-style rows below
    the main table, which use a different Job Code format).
  - Skips (with a warning) any row whose Job Code doesn't match the expected
    "##-USA-@@" pattern.
  - Skips (with a warning) any row with a Job Code but a blank Business Unit,
    rather than risk creating a blank Business Unit select option in Airtable.
  - Processes multiple files newest month -> oldest month, skipping any Job
    Code that's already in Airtable or was already added earlier in this
    same run (compared after trimming whitespace / normalizing nbsp so stray
    spaces don't create false "different job" duplicates).
  - Fetches existing Job Codes from Airtable even in --dry-run mode (a
    read-only call), so a dry-run preview's create/skip counts match what a
    real run would actually do -- not just duplicates within this run.
  - Warns if a Business Unit value isn't one of SYS/HT/EPC, and relies on the
    Airtable API's typecast=True option to auto-create the new Select option
    (this only needs the data.records:write scope, not schema.bases:write).

Requirements:
  - pip install openpyxl requests python-dotenv
  - An Airtable Personal Access Token (an existing token can be reused - see
    "Token scopes" below).

Token scopes (an existing token from the Next.js project can be reused):
  - data.records:read   (to fetch existing records for duplicate checking)
  - data.records:write  (to create new records + auto-create new Business
    Unit select options via typecast)
  - Those two scopes plus access to this Base are sufficient. No separate
    token needs to be issued just for this script.
"""

import argparse
import os
import re
import time
import unicodedata
from pathlib import Path

import openpyxl
import requests
from dotenv import load_dotenv

# ----------------------------------------------------------------------------
# Environment
# ----------------------------------------------------------------------------
PROJECT_ROOT = Path(__file__).resolve().parents[2]

if not load_dotenv(PROJECT_ROOT / ".env.local"):
    raise RuntimeError(f"Could not load {PROJECT_ROOT / '.env.local'}")

# ----------------------------------------------------------------------------
# Config
# ----------------------------------------------------------------------------
AIRTABLE_API_KEY = os.environ["AIRTABLE_API_KEY"]
AIRTABLE_BASE_ID = os.environ["AIRTABLE_BASE_ID"]

AIRTABLE_TABLE_NAME = "Jobs"  # table name or table ID (tblXXXXXXXX) both work

FIELD_JOB_CODE = "Job Code"
FIELD_JOB_NAME = "Job Name"
FIELD_BUSINESS_UNIT = "Business Unit"

DATA_START_ROW = 13
COL_BUSINESS_UNIT = 1
COL_JOB_CODE = 2
COL_JOB_NAME = 3

KNOWN_BUSINESS_UNITS = {"SYS", "HT", "EPC"}
JOB_CODE_PATTERN = re.compile(r"^\d{2}-USA-[A-Za-z0-9]+$")

AIRTABLE_API_ROOT = "https://api.airtable.com/v0"
BATCH_SIZE = 10  # Airtable allows up to 10 records per request
REQUEST_PAUSE_SEC = 0.25  # stay under Airtable's rate limit (5 req/sec)

# Multi-word/hyphenated phrases with inconsistent spacing/hyphenation/case in
# source spreadsheets. Matched case-insensitively regardless of spaces or
# hyphens, and normalized to one exact canonical spelling. Runs BEFORE
# title-casing so the canonical form is protected from being re-cased.
PHRASE_NORMALIZATIONS = [
    (re.compile(r"\bset[\s\-]+up\b", re.IGNORECASE), "Set Up"),
    (re.compile(r"\bhook[\s\-]+up\b", re.IGNORECASE), "Hook-up"),
]

# Known inconsistent spellings found in source spreadsheets, normalized to
# a single canonical form. Add new (wrong, right) pairs here as they're found.
NAME_NORMALIZATIONS = [
    ("T-PJT", "T PJT"),
]

# Obvious abbreviations/acronyms that must stay fully uppercase regardless of
# title-casing. Compared case-insensitively against each whitespace-separated
# word. Add new ones here as they're found (e.g. "GCS", "H2SO4").
ABBREVIATIONS = {"PJT", "OM", "FAB", "GCS", "CCSS", "SAS", "NBA", "MDU", "CUB",
                 "SCTA", "SPM", "SECAI", "SE&A", "B&D", "GA"}

# Words that must render in one specific exact casing, neither ALL CAPS
# (like an abbreviation) nor standard Title Case. Compared case-insensitively.
EXACT_CASE_WORDS = {"NAOH": "NaOH"}

# Already-canonical phrases (from PHRASE_NORMALIZATIONS above) that must be
# protected from further title-casing -- e.g. "Hook-up" should stay "Hook-up",
# not become "Hook-Up".
PROTECTED_PHRASES = {"Set Up", "Hook-up"}


# ----------------------------------------------------------------------------
# Utility functions
# ----------------------------------------------------------------------------
def title_case_preserving(job_name: str) -> str:
    words = job_name.split(" ")
    result = []
    skip_until = -1
    for i, word in enumerate(words):
        if i <= skip_until:
            continue
        matched_phrase = None
        for phrase in PROTECTED_PHRASES:
            phrase_words = phrase.split(" ")
            if words[i:i + len(phrase_words)] == phrase_words:
                matched_phrase = phrase
                skip_until = i + len(phrase_words) - 1
                break
        if matched_phrase:
            result.append(matched_phrase)
        elif any(ch.isdigit() for ch in word):
            # Alphanumeric codes/measurements (FAB2, U4, 2nm, 2Drum, 49HF...)
            # -- preserve exactly as extracted rather than risk mangling
            # embedded capitals.
            result.append(word)
        elif word.upper() in EXACT_CASE_WORDS:
            result.append(EXACT_CASE_WORDS[word.upper()])
        elif word.upper() in ABBREVIATIONS:
            result.append(word.upper())
        elif word == "":
            result.append(word)
        else:
            result.append(word[:1].upper() + word[1:].lower())
    return " ".join(result)


def normalize_job_name(job_name: str) -> str:
    # Exact 1:1 literal replacements (e.g. T-PJT -> T PJT)
    for wrong, right in NAME_NORMALIZATIONS:
        job_name = job_name.replace(wrong, right)

    # Case/spacing/hyphenation-insensitive phrase normalization
    for pattern, canonical in PHRASE_NORMALIZATIONS:
        job_name = pattern.sub(canonical, job_name)

    # Title-case everything else, protecting abbreviations/canonical phrases
    return title_case_preserving(job_name)


def normalize_str(value):
    """Trim leading/trailing whitespace, normalize nbsp, and collapse
    internal runs of whitespace into a single space."""
    if value is None:
        return ""
    text = str(value)
    text = unicodedata.normalize("NFKC", text)  # turns \xa0 (nbsp) etc. into a normal space
    text = text.replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def parse_year_month_from_filename(filename: str):
    """
    Extracts (year, month) from names like
    'Material Purchase Closing Report (2026-04).xlsx' or the
    underscore-sanitized 'Material_Purchase_Closing_Report__2026-04_.xlsx'.
    Works for any 20xx year (2024, 2025, 2026, ...), so this is reusable for
    past years' files without any code changes.
    """
    m = re.search(r"(20\d{2})[-_](\d{1,2})", filename)
    if not m:
        return None
    year, month = int(m.group(1)), int(m.group(2))
    if not (1 <= month <= 12):
        return None
    return year, month


def previous_year_month(year: int, month: int):
    if month == 1:
        return year - 1, 12
    return year, month - 1


def find_confirmed_sheet_name(workbook, year: int, month: int) -> str | None:
    """
    Finds the "previous month, confirmed" sheet inside a given year/month
    file. E.g. for a 2026-04 file, the previous month is 2026-03, so it
    looks for something matching '26 3월확정' (spacing around 월/확정 is
    flexible). The "추정" (estimate) sheet is excluded.
    """
    py, pm = previous_year_month(year, month)
    yy = py % 100
    # Accept both "25 12월확정" and "25 12월 확정" (with or without spaces)
    pattern = re.compile(rf"^\s*{yy:02d}\s*{pm}\s*월\s*확정\s*$")
    for name in workbook.sheetnames:
        if pattern.match(name.strip()):
            return name
    return None


def extract_rows_from_sheet(ws):
    """
    Starts at DATA_START_ROW and stops as soon as both columns A and B are
    empty in the same row.
    Returns: [{business_unit, job_code, job_name, row}]
    """
    rows = []
    row_idx = DATA_START_ROW
    max_row = ws.max_row
    while row_idx <= max_row:
        bu_raw = ws.cell(row=row_idx, column=COL_BUSINESS_UNIT).value
        code_raw = ws.cell(row=row_idx, column=COL_JOB_CODE).value
        name_raw = ws.cell(row=row_idx, column=COL_JOB_NAME).value

        bu = normalize_str(bu_raw)
        code = normalize_str(code_raw)
        name = normalize_str(name_raw)
        name = normalize_job_name(name)

        if not bu and not code:
            # Both A and B are empty -> end of the data section
            break

        if not code:
            row_idx += 1
            continue

        if not JOB_CODE_PATTERN.match(code):
            print(
                f"  [WARN] row {row_idx}: Job Code doesn't match the expected '##-USA-@@' pattern, skipping -> '{code}'")
            row_idx += 1
            continue

        if not bu:
            # A Job Code with no Business Unit -- skip rather than risk
            # creating a blank Business Unit select option via typecast.
            print(f"  [WARN] row {row_idx}: Business Unit is empty for Job Code '{code}', skipping row")
            row_idx += 1
            continue

        if bu not in KNOWN_BUSINESS_UNITS:
            print(f"  [INFO] row {row_idx}: new Business Unit value found -> '{bu}' (will be auto-added to Airtable)")

        rows.append({
            "business_unit": bu,
            "job_code": code,
            "job_name": name,
            "row": row_idx,
        })
        row_idx += 1

    return rows


def load_rows_from_file(filepath: Path) -> tuple[tuple[int, int] | None, list | None]:
    ym = parse_year_month_from_filename(filepath.name)
    if not ym:
        print(f"[SKIP] Could not extract year/month from filename: {filepath.name}")
        return None, None

    year, month = ym
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet_name = find_confirmed_sheet_name(wb, year, month)
    if not sheet_name:
        py, pm = previous_year_month(year, month)
        print(f"[SKIP] {filepath.name}: couldn't find a sheet matching '{py % 100:02d} {pm}월확정'. "
              f"Actual sheets in this file: {wb.sheetnames}")
        return (year, month), None

    ws = wb[sheet_name]
    print(f"[FILE] {filepath.name} -> using sheet '{sheet_name}'")
    rows = extract_rows_from_sheet(ws)
    return (year, month), rows


# ----------------------------------------------------------------------------
# Airtable API
# ----------------------------------------------------------------------------
class AirtableClient:
    def __init__(self, token: str, base_id: str, table_name: str):
        self.base_url = f"{AIRTABLE_API_ROOT}/{base_id}/{requests.utils.quote(table_name, safe='')}"
        self.headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
        }

    def fetch_existing_jobs(self):
        """Fetches every Job Code + Job Name already in the Jobs table.
        Returns a dict of {normalized_job_code: normalized_job_name}."""
        jobs = {}
        params: dict[str, str | int | list[str]] = {"fields[]": [FIELD_JOB_CODE, FIELD_JOB_NAME], "pageSize": 100}
        offset: str | None = None
        while True:
            if offset:
                params["offset"] = offset
            resp = requests.get(self.base_url, headers=self.headers, params=params)
            resp.raise_for_status()
            data = resp.json()
            for record in data.get("records", []):
                fields = record.get("fields", {})
                code = fields.get(FIELD_JOB_CODE)
                name = fields.get(FIELD_JOB_NAME)
                if code:
                    jobs[normalize_str(code)] = normalize_job_name(normalize_str(name or ""))
            offset = data.get("offset")
            if not offset:
                break
            time.sleep(REQUEST_PAUSE_SEC)
        return jobs

    def create_records(self, records, dry_run=False):
        """records: [{business_unit, job_code, job_name}] -> batch-creates
        them in Airtable."""
        created = 0
        for i in range(0, len(records), BATCH_SIZE):
            batch = records[i:i + BATCH_SIZE]
            payload = {
                "typecast": True,  # allow auto-creation of new Business Unit select options
                "records": [
                    {
                        "fields": {
                            FIELD_JOB_CODE: r["job_code"],
                            FIELD_JOB_NAME: r["job_name"],
                            FIELD_BUSINESS_UNIT: r["business_unit"],
                        }
                    }
                    for r in batch
                ],
            }
            if dry_run:
                for r in batch:
                    print(f"    (dry-run) would create -> {r['job_code']} | {r['job_name']} | {r['business_unit']}")
                created += len(batch)
                continue

            resp = requests.post(self.base_url, headers=self.headers, json=payload)
            if resp.status_code >= 300:
                print(f"    [ERROR] Airtable create failed ({resp.status_code}): {resp.text}")
                resp.raise_for_status()
            created += len(batch)
            time.sleep(REQUEST_PAUSE_SEC)
        return created


# ----------------------------------------------------------------------------
# Main logic
# ----------------------------------------------------------------------------
def collect_target_files(folder: str | None = None, files: list | None = None):
    if files:
        paths = [Path(f) for f in files]
    elif folder:
        folder_path = Path(folder)
        paths = sorted(folder_path.glob("*Material*Purchase*Closing*Report*.xlsx"))
    else:
        raise ValueError("You must specify either --folder or --files.")

    dated = []
    for p in paths:
        ym = parse_year_month_from_filename(p.name)
        if ym:
            dated.append((ym, p))
        else:
            print(f"[SKIP] Could not recognize year/month in filename: {p.name}")

    # Sort newest month -> oldest month (reverse chronological processing order)
    dated.sort(key=lambda x: x[0], reverse=True)
    return dated


def main():
    parser = argparse.ArgumentParser(description="Sync Material Purchase Closing Report data into Airtable Jobs")
    parser.add_argument("--folder", help="Folder containing the Excel files")
    parser.add_argument("--files", nargs="*", help="Specific Excel file paths to process (alternative to --folder)")
    parser.add_argument("--base-id", default=AIRTABLE_BASE_ID, help="Airtable Base ID")
    parser.add_argument("--table", default=AIRTABLE_TABLE_NAME, help="Airtable table name/ID")
    parser.add_argument("--dry-run", action="store_true", help="Preview only, without writing to Airtable")
    args = parser.parse_args()

    token = AIRTABLE_API_KEY

    targets = collect_target_files(folder=args.folder, files=args.files)
    if not targets:
        print("No files to process.")
        return

    print("Processing order (newest month -> oldest month):")
    for (y, m), p in targets:
        print(f"  - {y}-{m:02d}: {p.name}")
    print()

    # Fetch existing Job Codes regardless of --dry-run: this is a read-only
    # call, and skipping it during dry-run would make the preview's
    # create/skip counts diverge from what a real run would actually do.
    client = AirtableClient(token, args.base_id, args.table)
    print("Fetching existing Job Codes from Airtable...")
    existing_jobs = client.fetch_existing_jobs()
    print(f"Found {len(existing_jobs)} existing Job Codes.\n")

    seen_this_run = dict(existing_jobs)  # {code: name}
    total_created = 0
    total_skipped_dup = 0
    total_name_mismatches = 0

    for (_year, _month), filepath in targets:
        _, rows = load_rows_from_file(filepath)
        if rows is None:
            continue

        to_create = []
        for r in rows:
            code_key = r["job_code"]
            normalized_name = normalize_job_name(r["job_name"])

            if code_key in seen_this_run:
                existing_name = seen_this_run[code_key]
                if existing_name and normalized_name != existing_name:
                    print(f"  [WARN] Job Code '{code_key}' already exists with a different "
                          f"Job Name -- existing: '{existing_name}' | this file: '{normalized_name}'")
                    total_name_mismatches += 1
                total_skipped_dup += 1
                continue

            seen_this_run[code_key] = normalized_name
            to_create.append(r)

        print(f"  -> {len(to_create)} new / {len(rows) - len(to_create)} skipped as duplicates")

        if to_create:
            created = client.create_records(to_create, dry_run=args.dry_run)
            total_created += created
        print()

    print("=" * 50)
    print(f"Done: {total_created} record(s) created, {total_skipped_dup} skipped as duplicates "
          f"({total_name_mismatches} of those had a Job Name mismatch worth reviewing)")


if __name__ == "__main__":
    main()
